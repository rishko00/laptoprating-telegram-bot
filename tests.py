import kursach as k
import unittest
from openpyxl import Workbook, load_workbook


class Test(unittest.TestCase):

    def test_sql_to_xlsx(self):
        k.cur.execute('SELECT * FROM Laptop')
        rows = k.cur.fetchall()
        wb = load_workbook('db.xlsx')
        ws = wb.active
        self.assertEqual(ws.max_row, len(rows))
        self.assertEqual(ws.max_column, len(rows[0]))
        for i in range(1, ws.max_row):
            for j in range(1, ws.max_column):
                self.assertEqual(ws.cell(row = i, column = j).value, str(rows[i - 1][j - 1]))
    

    def test_xlsx_to_sql(self):
        k.cur.execute('SELECT * FROM Laptop')
        rows = k.cur.fetchall()
        wb = load_workbook('db_from.xlsx')
        ws = wb.active
        for i in range(1, ws.max_row):
            for j in range(1, ws.max_column):
                self.assertEqual(ws.cell(row = i, column = j).value, str(rows[len(rows) - i - 1][j])) 

    def test_add(self):
        k.cur.execute('SELECT * FROM Laptop')
        rows = k.cur.fetchall()
        data = [
            ['Asus', 'LLKM', 'Info', 11123],
            ['Acer', 'LM', 'Info', 12323],
            ['HP', 'M9776', 'Info', 11123]
        ]
        for i in range(len(data)):
            k.cur.execute('INSERT INTO Laptop (brand, model, info, price) VALUES ("{0}", "{1}", "{2}", {3})'.format(data[i][0], data[i][1], data[i][2], data[i][3]))
            k.cur.execute('SELECT * FROM Laptop')
            rows2 = k.cur.fetchall()
            k.cur.connection.rollback()
            self.assertLess(len(rows), len(rows2))

    def test_delete(self):
        k.cur.execute('SELECT * FROM Laptop')
        rows = k.cur.fetchall()

        data = [
            ['Asus', 'LLKM'],
            ['Acer', 'LM'],
            ['HP', 'M9776'],
            ['Dell', '87nY']
        ]
        for i in range(len(data)):
            k.cur.execute('DELETE FROM Laptop WHERE brand = "{0}" and model = "{1}"'.format(data[i][0], data[i][1]))
            k.cur.execute('SELECT * FROM Laptop')
            rows2 = k.cur.fetchall()
            k.cur.connection.rollback()
            self.assertEqual(len(rows), len(rows2))

        data_true = [
            ['Asus', 'Assssd'],
            ['Asus', 'LINJ'],
            ['HP', 'HDGHJS']
        ]
        for i in range(len(data_true)):
            k.cur.execute('DELETE FROM Laptop WHERE brand = "{0}" and model = "{1}"'.format(data_true[i][0], data_true[i][1]))
            k.cur.execute('SELECT * FROM Laptop')
            rows2 = k.cur.fetchall()
            k.cur.connection.rollback()
            self.assertLess(len(rows2), len(rows))

            
if __name__ == '__main__':
    unittest.main()
