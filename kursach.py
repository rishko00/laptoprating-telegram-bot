import pymysql
import time
import telebot
import xlrd
import csv
import sys
from openpyxl import Workbook, load_workbook
from decimal import Decimal

connection = pymysql.connect(
    host = 'localhost',
    user = 'root',
    password = 'Pass_1111',
    db = 'Laptops',
    charset = 'utf8mb4'
)
cur = connection.cursor()

bot = telebot.TeleBot(os.getenv('API_KEY'))

def update_keyboard():
    brands = []
    cur.execute('SELECT brand, model FROM Laptop ORDER BY brand, model')
    rows = cur.fetchall()
    keyboard1 = telebot.types.ReplyKeyboardMarkup()
    for row in rows:
        brands.append(row[0] + ' ' + row[1])
    keyboard1.add(*brands, row_width=1)
    return keyboard1


def inline_keyboard():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton(text='Поставити оцінку', callback_data='rate'))
    return keyboard


@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, 'Список ноутбуків:', reply_markup = update_keyboard())


@bot.message_handler(commands=['sql_to_xlsx'])
def start_message(message):
    query = "SELECT * FROM Laptop"
    cur.execute(query)
    result = cur.fetchall()
    f = open('dbdump.csv', 'w', encoding='utf-8')
    c = csv.writer(f)
    for i in result:
        c.writerow(i)
    f.close()

    wb = Workbook()
    ws = wb.active
    with open('dbdump.csv', 'r', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            ws.append(row)
        wb.save('db.xlsx')

    bot.send_message(message.chat.id, 'Успішно виконано!', reply_markup = update_keyboard())


@bot.message_handler(commands=['xlsx_to_sql'])
def start_message(message):
    wb = load_workbook('db_from.xlsx')
    result = []
    sheets = wb.sheetnames

    for sheet in sheets:
        ws = wb[sheet] 
        for row in ws.values:
            result.append(row)
    wb.save('db_from.xlsx')

    f = open('dbdump_from.csv', 'w', encoding='utf-8')
    c = csv.writer(f)
    for i in result:
        c.writerow(i)
    f.close()

    with open('dbdump_from.csv', 'r', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            query = 'SELECT * FROM Laptop where brand = "{0}" and model = "{1}"'.format(row[0], row[1])
            cur.execute(query)
            if cur.fetchall():
                bot.send_message(message.chat.id, 'Ноутбук вже є  в базі даних!')
                continue
                
            query2 = 'INSERT INTO Laptop(brand, model, info, price) VALUES("{0}", "{1}", "{2}", {3})'.format(row[0], row[1], row[2], row[3])
            cur.execute(query2)

    connection.commit()
    bot.send_message(message.chat.id, 'Успішно виконано!', reply_markup = update_keyboard())


class Laptop:
    def __init__(self, brand, model):
        self.brand = brand
        self.model = model
        self.info = ''
        self.price = 0

curr_laptop = []
        
@bot.message_handler(commands=['add'])
def add_laptop(message):
    msg = bot.send_message(message.chat.id, 'Введіть виробника і модель ноутбука:')
    bot.register_next_step_handler(msg, laptop_model)

def laptop_model(message):
    data_laptop = message.text.split(' ', 1)
    global laptop 
    laptop = Laptop('','')
    try:
        laptop.brand = data_laptop[0]
        laptop.model = data_laptop[1]
        cur.execute('SELECT * FROM Laptop WHERE brand = "{0}" and model = "{1}"'.format(laptop.brand, laptop.model))
        rows = cur.fetchall()
        if len(rows) > 0:
            curr_laptop.extend(data_laptop)
            bot.send_message(message.chat.id, 'Такий ноутбук вже є в базі даних!')
            for row in rows:
                bot.send_message(message.chat.id, 
                    'Виробник: {0}\nМодель: {1}\nХарактеристики:\n{2} \nЦіна: {3}\nРейтинг: {4}\nКількість оцінок: {5}'
                    .format(row[1], row[2], row[3], row[4], row[5], row[6]), reply_markup = inline_keyboard())
        else:
            msg = bot.send_message(message.chat.id, 'Введіть характеристики ноутбука:')
            bot.register_next_step_handler(msg, laptop_info)
    
    except:
        bot.send_message(message.chat.id, 'Введіть коректні дані!')


def laptop_info(message):
    laptop.info = message.text
    msg = bot.send_message(message.chat.id, 'Введіть ціну:')
    bot.register_next_step_handler(msg, laptop_submit)

def laptop_submit(message):
    try:
        laptop.price = int(message.text)
        cur.execute('SELECT * FROM Laptop WHERE brand = "{0}" and model = "{1}"'.format(laptop.brand, laptop.model))
        if len(cur.fetchall()) == 1:
            bot.send_message(message.chat.id, 'Такий ноутбук вже є в базі даних!')
        else:
            cur.execute('INSERT INTO Laptop (brand, model, info, price) VALUES ("{0}", "{1}", "{2}", {3})'.format(laptop.brand, laptop.model, laptop.info, laptop.price))
            msg = bot.send_message(message.chat.id, 'Ноутбук додано до бази даних!', reply_markup = update_keyboard())
        connection.commit()
    
    except:
        bot.send_message(message.chat.id, 'Введіть числове значення ціни!')


@bot.message_handler(commands=['delete'])
def start_message(message):
    msg = bot.send_message(message.chat.id, 'Введіть виробника і модель ноутбука:')
    bot.register_next_step_handler(msg, delete_laptop)

def delete_laptop(message):
    try:
        laptop_name = message.text.split(' ', 1)
        laptop = Laptop(laptop_name[0], laptop_name[1])
        cur.execute('SELECT * FROM Laptop WHERE brand = "{0}" and model = "{1}"'.format(laptop.brand, laptop.model))
        if len(cur.fetchall()) == 0:
            bot.send_message(message.chat.id, 'Такого ноутбука немає в базі даних!')
        else:
            cur.execute('DELETE FROM Laptop WHERE brand = "{0}" and model = "{1}"'.format(laptop.brand, laptop.model))
            bot.send_message(message.chat.id, 'Ноутбук успішно видалено!', reply_markup = update_keyboard())
        connection.commit()
    except:
        bot.send_message(message.chat.id, 'Введіть коректні дані!')


@bot.message_handler(content_types=['text'])
def send_text(message):
    try:
        find_laptop = message.text.split(' ', 1)
        cur.execute('SELECT * FROM Laptop WHERE brand = "{0}" AND model = "{1}"'.format(find_laptop[0], find_laptop[1]))
        rows = cur.fetchall()
        if rows:
            curr_laptop.extend(find_laptop)
            for row in rows:
                curr_laptop[2] = row[0]
                bot.send_message(message.chat.id, 
                'Виробник: {0}\nМодель: {1}\nХарактеристики:\n{2} \nЦіна: {3}\nРейтинг: {4}\nКількість оцінок: {5}'
                .format(row[1], row[2], row[3], row[4], row[5], row[6]), reply_markup = inline_keyboard())
    
    except:
        bot.send_message(message.chat.id, 'Введіть коректні дані!')
         

@bot.callback_query_handler(func=lambda call: True)
def query_handler(call):
    keyboard = telebot.types.InlineKeyboardMarkup()
    if call.data == 'rate':
        cur.execute('SELECT * FROM Rating WHERE user_id = {0} AND laptop_id = {1}'.format(call.from_user.id, curr_laptop[2]))
        rate_fromuser = cur.fetchall()
        if len(rate_fromuser) == 1:
            for row in rate_fromuser:
                bot.send_message(call.message.chat.id, 'Ви вже оцінювали цей ноутбук, ваша оцінка: {0}'.format(row[3]))
        else:
            keyboard.add(telebot.types.InlineKeyboardButton(text='1', callback_data='1'))
            keyboard.add(telebot.types.InlineKeyboardButton(text='2', callback_data='2'))
            keyboard.add(telebot.types.InlineKeyboardButton(text='3', callback_data='3'))
            keyboard.add(telebot.types.InlineKeyboardButton(text='4', callback_data='4'))
            keyboard.add(telebot.types.InlineKeyboardButton(text='5', callback_data='5'))
            bot.send_message(call.message.chat.id, 'Rate', reply_markup=keyboard)
    elif int(call.data) in [1, 2, 3, 4, 5]:
        cur.execute('INSERT INTO Rating(user_id, laptop_id, rate) VALUES("{0}", {1}, {2})'.format(call.from_user.id, curr_laptop[2], int(call.data)))
        cur.execute('SELECT rating, r_count, info FROM Laptop WHERE brand = "{0}" AND model = "{1}"'.format(curr_laptop[0], curr_laptop[1]))
        rows = cur.fetchall()
        count = 0
        rate = 0.0
        for row in rows:
            count = row[1] + 1
            rate = int(call.data) if row[0] == 0 else (row[0] + int(call.data)) / count
        cur.execute('UPDATE Laptop SET rating="{0}", r_count={1} WHERE brand = "{2}" AND model = "{3}"'.format(Decimal(rate), count, curr_laptop[0], curr_laptop[1]))
        bot.send_message(call.message.chat.id, 'Ваша оцінка: {0} \nДякуємо за відгук!'.format(call.data))
    connection.commit()
    

bot.polling()

